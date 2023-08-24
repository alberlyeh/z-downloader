import os
from PyPDF2 import PdfReader
import re
import openpyxl
from colorama import init, Fore

init(autoreset=True)  # Inizializza colorama

def estrai_testo_da_pdf(pdf_file):
    testo = ""
    pdf_reader = PdfReader(pdf_file)
    for pagina in pdf_reader.pages:
        testo += pagina.extract_text()
    return testo

pdf_files = [f for f in os.listdir() if f.lower().endswith(".pdf") and f.startswith("Libro_unico")]

parola_chiave = " €"

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Dati"

sheet.cell(row=1, column=1, value="DATA")
sheet.cell(row=1, column=2, value="NETTO DEL MESE")

row_number = 2
data_list = []

for pdf_file in pdf_files:
    match = re.search(r'(\d{2}-\d{4})', pdf_file)
    if match:
        data = match.group(1)
    else:
        data = "Data non trovata"
    
    testo_pdf = estrai_testo_da_pdf(pdf_file)
    
    if parola_chiave in testo_pdf:
        index = testo_pdf.find(parola_chiave)
        euro = ""
        i = index - 1
        while i >= 0 and (testo_pdf[i].isdigit() or testo_pdf[i] == ',' or testo_pdf[i] == '.'):
            euro = testo_pdf[i] + euro
            i -= 1
        euro = euro.strip()
        if euro:
            sheet.cell(row=row_number, column=1, value=data)
            sheet.cell(row=row_number, column=2, value=euro)
            data_list.append({"DATA": data, "NETTO DEL MESE": euro})
            row_number += 1
            print(f"File: {Fore.GREEN}{pdf_file}{Fore.RESET}, Data: {Fore.CYAN}{data}{Fore.RESET}, Importo: {Fore.GREEN}{euro}{Fore.RESET}")
    else:
        print(f"File: {Fore.RED}{pdf_file}{Fore.RESET}, Data: {Fore.CYAN}{data}{Fore.RESET}, {Fore.RED}Nessun importo trovato{Fore.RESET}")

workbook.save("output.xls")
print("Dati salvati in output.xls")
